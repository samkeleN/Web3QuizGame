"""
File parsing utilities for the AI Project Analyzer.
"""

import os
import re
import logging
import csv
from typing import List, Dict, Any
from openpyxl import load_workbook  # Only for Excel support

logger = logging.getLogger(__name__)

def parse_input_file(file_path: str) -> List[str]:
    """
    Parse an Excel or CSV file to extract GitHub repository URLs.

    Args:
        file_path: Path to the Excel or CSV file containing GitHub URLs.

    Returns:
        List of GitHub repository URLs.

    Raises:
        ValueError: If the file type is not supported or no GitHub URLs are found.
        FileNotFoundError: If the file does not exist.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    # Determine file type from extension
    file_ext = os.path.splitext(file_path)[1].lower()

    try:
        data = []
        columns = []
        
        if file_ext == ".csv":
            with open(file_path, mode='r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                columns = reader.fieldnames or []
                data = [row for row in reader]
                
        elif file_ext in [".xlsx", ".xls"]:
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb.active
            columns = [cell.value for cell in sheet[1]]  # First row as headers
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(columns, row)))
        else:
            raise ValueError(
                f"Unsupported file type: {file_ext}. Please provide a CSV or Excel file."
            )

        # Find columns that might contain GitHub URLs
        github_columns = find_github_columns(columns)

        if not github_columns:
            raise ValueError(
                "No GitHub URL columns found in the file. "
                "Please ensure your file has a column with 'github' or 'Github' in the name."
            )

        # Extract URLs from the identified columns
        urls = extract_github_urls(data, github_columns)

        if not urls:
            raise ValueError("No valid GitHub repository URLs found in the file.")

        return urls

    except Exception as e:
        logger.error(f"Error parsing file {file_path}: {str(e)}")
        raise


def find_github_columns(columns: List[str]) -> List[str]:
    """
    Find columns that might contain GitHub URLs.

    Args:
        columns: List of column names to search

    Returns:
        List of column names that might contain GitHub URLs.
    """
    github_columns = []
    for col in columns:
        if col and re.search(r"(?i)github|github url", str(col)):
            github_columns.append(col)
    return github_columns


def extract_github_urls(data: List[Dict[str, Any]], columns: List[str]) -> List[str]:
    """
    Extract GitHub repository URLs from specified columns in the data.

    Args:
        data: List of dictionaries containing the row data
        columns: List of column names to extract GitHub URLs from

    Returns:
        List of valid GitHub repository URLs.
    """
    github_urls = []
    github_pattern = re.compile(r"https?://(?:www\.)?github\.com/[\w.-]+/[\w.-]+/?")

    for col in columns:
        for row in data:
            value = str(row.get(col, "")).strip()
            if match := github_pattern.search(value):
                url = match.group(0)
                # Clean up URL
                if url.endswith(")") and "(" not in url:
                    url = url[:-1]
                if url.endswith("/"):
                    url = url[:-1]
                if url not in github_urls:
                    github_urls.append(url)

    return github_urls


def validate_github_url(url: str) -> bool:
    """
    Validate if a URL is a proper GitHub repository URL.

    Args:
        url: URL to validate.

    Returns:
        True if the URL is a valid GitHub repository URL, False otherwise.
    """
    github_repo_pattern = re.compile(r"^https?://(?:www\.)?github\.com/[\w.-]+/[\w.-]+/?$")
    return bool(github_repo_pattern.match(url))